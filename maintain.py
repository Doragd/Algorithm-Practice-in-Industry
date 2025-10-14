import re
import ast
import os
import random
import json
import argparse
import openpyxl
import requests
import datetime

# 支持多个飞书URL，使用逗号分隔
FEISHU_URLS = os.environ.get("FEISHU_URL", "").split(',')
# 去除空字符串和空格
FEISHU_URLS = [url.strip() for url in FEISHU_URLS if url.strip()]

def set_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--issue", "-i", type=str, help="input issue", required=True)
    args = parser.parse_args()
    return args


def parse_issue(issue):
    try:
        issue = issue.replace("\\n", "").replace("\\r", "")
        info = ast.literal_eval(issue)
        print(info)
        assert isinstance(info, list)
        assert len(info) > 0 and info[0].get("公司") and info[0].get("内容") and info[0].get("标签") and info[0].get("时间") and info[0].get("链接")
    except:
        raise Exception("[-] Wrong input!")
    return info

def write_item(wb, item):
    columns = ["公司", "内容", "标签", "时间"]
    values = [item[column] for column in columns]
    ws = wb.active
    ws.append(values)
    ws.cell(row=ws.max_row, column=2).style = 'Hyperlink'
    ws.cell(row=ws.max_row, column=2).hyperlink = item["链接"]
    print("[+] Write item: {}".format(values))

def update_excel(args):
    print("[+] Add new items into excel...")
    wb = openpyxl.load_workbook("source.xlsx")
    info = parse_issue(args.issue)
    for item in info:
        assert item.keys() == {"公司", "内容", "标签", "时间", "链接"}
        write_item(wb, item)
    wb.save("source.xlsx")
    print("[+] Add items Done!")

def update_readme(args, info=None):
    print("[+] Add new items into readme...")
    if info is None:
        info = parse_issue(args.issue)
    with open("README.md", "r", encoding="utf-8") as f:
        lines = f.readlines()

    # 查找表格位置
    table_title = "## 大厂实践文章"
    table_index = None
    for i, line in enumerate(lines):
        if line.strip() == table_title:
            table_index = i
            break

    # 如果找到表格位置，则更新表格
    if table_index is not None:
        j = table_index + 2
        while j < len(lines) and not re.search(r"\|(\s-+\s\|)+", lines[j]):
            j += 1
        index = j + 1
        newlines = []
        for item in info:
            newline = "".join("| {} | [{}]({}) | {} | {} |\n".format(
                item["公司"], item["内容"], item["链接"], item["标签"], item["时间"]))
            newlines.append(newline)
        lines = lines[:index] + newlines + lines[index:]
        with open("README.md", "w", encoding="utf-8") as f:
            f.writelines(lines)
        print("[+] Add items Done!")
    else:
        print("[-] Table not found!")

def update_message(args):
    infos = parse_issue(args.issue)
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    title = f"🌸大厂实践文章自动更新@{today}"
    content = []
    for info in infos:
        emoji = random.choice("🌱🌿🍀🪴🎋🍃🪷🌸✨")
        meta_info = f"🥹 {info['公司']}  📆 {info['时间']}  🍉 {info['标签']}"
        info_title = f"✅ {info['内容']}"
        line_len = max(len(meta_info), len(info_title))
        sepline = emoji * line_len
        content.append(
            [{
                "tag": "text",
                "text": ""
            }]
        )
        # content.append(
        #     [{
        #         "tag": "text",
        #         "text": sepline
        #     }]
        # )
        # content.append(
        #     [{
        #         "tag": "text",
        #         "text": ""
        #     }]
        # )
        content.append(
            [{
                "tag": "text",
                "text": meta_info
            }]
        )
        content.append(
            [{
                "tag": "text",
                "text": ""
            }]
        )
        content.append(
            [{
                "tag": "a",
                "text": info_title,
                "href": f"{info['链接']}"
            }]
        )
        content.append(
            [{
                "tag": "text",
                "text": ""
            }]
        )
    content.append(
        [{
            "tag": "text",
            "text": "-----",
        }]
    )
    content.append(
        [{
            "tag": "a",
            "text": "➡️ 点击查看完整大厂实践文章列表",
            "href": "https://github.com/Doragd/Algorithm-Practice-in-Industry"
        }]
    )
    send_feishu_message(title, content)

def send_feishu_message(title, content, urls=None):
    # 如果没有指定URL列表，使用默认的FEISHU_URLS
    if urls is None:
        urls = FEISHU_URLS
    
    # 如果没有有效的飞书URL，直接返回
    if not urls:
        print("⚠️ 没有有效的飞书URL，跳过发送消息")
        return
    
    raw_data = {
        "msg_type": "post",
        "content": {
            "post": {
                "zh_cn": {
                    "title": title,
                    "content": content
                }
            }
        }
    }  
    body = json.dumps(raw_data)
    headers = {"Content-Type":"application/json"}
    
    # 向每个飞书URL发送消息
    for idx, url in enumerate(urls):
        try:
            ret = requests.post(url=url, data=body, headers=headers, timeout=10)
            print(f"✉️ 飞书推送[{idx+1}/{len(urls)}]返回: {ret.text}")
        except Exception as e:
            print(f"❌ 飞书推送[{idx+1}/{len(urls)}]失败: {e}")



def main():
    args = set_args()
    update_excel(args)
    update_readme(args)
    update_message(args)

if __name__ == "__main__":
    main()
